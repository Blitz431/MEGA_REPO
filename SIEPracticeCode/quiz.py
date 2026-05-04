"""
SIE Exam Quizzer
----------------
Reads questions from "QUESTION BANK AND TRACKER.xlsx" and quizzes the user
via a CLI. Writes RIGHT / WRONG / # times asked back to the Excel file.

Usage:  python quiz.py
"""

import os
import sys
import csv
import json
import time
import random
import datetime
import subprocess

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is not installed. Run:  pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EXCEL_PATH     = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "QUESTION BANK AND TRACKER.xlsx")
LOG_PATH       = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "session_log.csv")
BOOKMARKS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "bookmarks.json")

SECTION_MAP = {
    "1": {
        "name": "Knowledge of Capital Markets",
        "subsections": ["1.1", "1.2", "1.3"],
        "sub_names": {
            "1.1": "Regulatory, Agencies, & Market",
            "1.2": "Market Structure",
            "1.3": "Economic Factors",
        },
    },
    "2": {
        "name": "Understanding Products and Their Risks",
        "subsections": ["2.1", "2.2", "2.3", "2.4", "2.5"],
        "sub_names": {
            "2.1": "Equity Securities",
            "2.2": "Debt Securities",
            "2.3": "Packaged Products",
            "2.4": "Options",
            "2.5": "Alt Investments & Other Products",
        },
    },
    "3": {
        "name": "Understanding Trading, Customer Accounts & Prohibited Activities",
        "subsections": ["3.1", "3.2", "3.3"],
        "sub_names": {
            "3.1": "Trading Securities",
            "3.2": "Customer Accounts & Compliance",
            "3.3": "Prohibited Activities",
        },
    },
    "4": {
        "name": "Overview of the Regulatory Framework",
        "subsections": ["4.1", "4.2", "4.3"],
        "sub_names": {
            "4.1": "SROs and Regulatory Bodies",
            "4.2": "Registration & Licensing",
            "4.3": "Key Regulations & Rules",
        },
    },
}

# Real SIE exam distribution
EXAM_DISTRIBUTION = {"1": 12, "2": 33, "3": 23, "4": 7}

PASS_THRESHOLD = 0.72  # 72% to pass the SIE

DIVIDER  = "=" * 62
THIN_DIV = "-" * 62

# ---------------------------------------------------------------------------
# Help text
# ---------------------------------------------------------------------------

HELP_TEXT = """
╔══════════════════════════════════════════════════════════════╗
║                    SIE QUIZZER — HELP                       ║
╠══════════════════════════════════════════════════════════════╣
║  MODES                                                      ║
║  F  Full Exam    75 questions, real SIE distribution        ║
║  S  Study Mode   pick sections/subsections + target count   ║
║  W  Weak Areas   only questions you've answered wrong       ║
║  B  Bookmarks    re-quiz your bookmarked questions          ║
║  V  View History show your past session scores              ║
╠══════════════════════════════════════════════════════════════╣
║  CHECKLIST MENUS                                            ║
║  <number>  toggle that item on/off                          ║
║  A         select all                                       ║
║  C         clear all                                        ║
║  Enter     confirm selection                                ║
╠══════════════════════════════════════════════════════════════╣
║  ANSWERING QUESTIONS                                        ║
║  1–4       pick your answer (choices are shuffled each time)║
║  IDK       skip — marks wrong, adds to review list          ║
║  C         copy question + answer to clipboard              ║
║  B         bookmark (in Exam Mode, type B at answer prompt) ║
║  END TEST  end Full Exam early and go straight to review    ║
║  H or ?    show this help screen                            ║
║  Q         quit to main menu                                ║
║  Ctrl+C    exit the program                                 ║
╠══════════════════════════════════════════════════════════════╣
║  CONFIDENCE (shown after correct answers in study modes)    ║
║  1  Guessed    — shown again like a wrong answer (penalized)║
║  2  Unsure     — shown more often than average              ║
║  3  Kinda sure — shown at normal frequency                  ║
║  4  Sure       — shown less often (you know it cold)        ║
║  Enter         — skip, treated the same as kinda sure (3)   ║
║  Add ,B to combine rating + bookmark  (e.g.  3,B  or  B,2) ║
╠══════════════════════════════════════════════════════════════╣
║  TRACKING                                                   ║
║  RIGHT / WRONG / # times asked are saved to Excel after     ║
║  each answer. Close Excel before running to allow saving.   ║
║  Column K in Excel = explanation text (fill in anytime).    ║
╚══════════════════════════════════════════════════════════════╝
"""

# ---------------------------------------------------------------------------
# Data layer
# ---------------------------------------------------------------------------

def load_all_questions() -> list:
    """Load every question from the Excel file. Returns list of dicts."""
    if not os.path.exists(EXCEL_PATH):
        print(f"\nERROR: Excel file not found at:\n  {EXCEL_PATH}", flush=True)
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    except Exception as exc:
        print(f"\nERROR: Could not open Excel file: {exc}", flush=True)
        sys.exit(1)

    ws = wb.active
    questions = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 10:
            continue

        section_raw, question, choice_a, choice_b, choice_c, choice_d, \
            correct, times, right, wrong = (row[i] for i in range(10))

        # Column K (index 10): optional explanation — user can fill this in anytime
        explanation = str(row[10]).strip() if len(row) > 10 and row[10] is not None else ""
        # Column L (index 11): lucky guesses
        lucky = int(row[11]) if len(row) > 11 and row[11] is not None else 0
        # Column M (index 12): cumulative confidence points (2=kinda sure baseline per right answer)
        conf_sum = int(row[12]) if len(row) > 12 and row[12] is not None else 0

        if section_raw is None:
            continue

        # Normalize section to "X.X" string (it may be stored as a float)
        try:
            section_str = f"{float(section_raw):.1f}"
        except (ValueError, TypeError):
            section_str = str(section_raw).strip()

        # Skip rows with missing data
        if any(v is None for v in [question, choice_a, choice_b, choice_c, choice_d, correct]):
            print(f"  [Warning] Skipping incomplete row {row_idx}", flush=True)
            continue

        correct = str(correct).strip().upper()
        if correct not in {"A", "B", "C", "D"}:
            print(f"  [Warning] Row {row_idx} has invalid correct answer '{correct}' — skipping", flush=True)
            continue

        questions.append({
            "row":         row_idx,
            "section":     section_str,
            "question":    str(question),
            "choices":     {
                "A": str(choice_a),
                "B": str(choice_b),
                "C": str(choice_c),
                "D": str(choice_d),
            },
            "correct":     correct,
            "explanation": explanation,
            "times_asked": int(times) if times is not None else 0,
            "right":       int(right) if right  is not None else 0,
            "wrong":       int(wrong) if wrong  is not None else 0,
            "lucky":       lucky,
            "conf_sum":    conf_sum,
        })

    wb.close()
    return questions


def save_result(row: int, correct: bool, lucky: bool = False, conf_pts: int = 0) -> None:
    """Increment tracking columns in the Excel file for the given row."""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        def inc(col):
            v = ws.cell(row=row, column=col).value
            ws.cell(row=row, column=col).value = (int(v) if v is not None else 0) + 1

        def add(col, n):
            v = ws.cell(row=row, column=col).value
            ws.cell(row=row, column=col).value = (int(v) if v is not None else 0) + n

        if ws.cell(row=1, column=12).value is None:
            ws.cell(row=1, column=12).value = "LUCKY"
        if ws.cell(row=1, column=13).value is None:
            ws.cell(row=1, column=13).value = "CONF_SUM"

        inc(8)  # # Of times asked
        if correct:
            inc(9)   # RIGHT
        else:
            inc(10)  # WRONG
        if lucky:
            inc(12)  # LUCKY
        if conf_pts > 0:
            add(13, conf_pts)  # CONF_SUM

        wb.save(EXCEL_PATH)
        wb.close()
    except PermissionError:
        print("\n  [Warning] Could not save to Excel — close the file and re-run to save stats.")
    except Exception as exc:
        print(f"\n  [Warning] Could not save result: {exc}")


def append_session_log(mode: str, sections: str, target: int,
                       correct: int, total: int) -> None:
    """Append one line to session_log.csv."""
    pct = (correct / total * 100) if total > 0 else 0.0
    write_header = not os.path.exists(LOG_PATH)
    try:
        with open(LOG_PATH, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if write_header:
                writer.writerow(["date", "mode", "sections", "target",
                                 "correct", "total", "pct"])
            writer.writerow([
                datetime.date.today().isoformat(),
                mode, sections, target, correct, total,
                f"{pct:.1f}",
            ])
    except Exception:
        pass  # Don't crash if log write fails


def load_bookmarks() -> set:
    """Load bookmarked row numbers from bookmarks.json."""
    if not os.path.exists(BOOKMARKS_PATH):
        return set()
    try:
        with open(BOOKMARKS_PATH, "r", encoding="utf-8") as f:
            return set(json.load(f))
    except Exception:
        return set()


def save_bookmarks(bookmarks: set) -> None:
    """Save bookmarked row numbers to bookmarks.json."""
    try:
        with open(BOOKMARKS_PATH, "w", encoding="utf-8") as f:
            json.dump(sorted(bookmarks), f)
    except Exception as exc:
        print(f"  [Warning] Could not save bookmarks: {exc}")

# ---------------------------------------------------------------------------
# Adaptive weighting
# ---------------------------------------------------------------------------

def question_weight(q: dict) -> float:
    """Higher weight = shown more often (weak questions prioritized).

    Confidence ratings accumulate in conf_sum (2 pts = kinda-sure baseline):
      rating 1 (guessed)   → 0 pts, also counted as wrong via lucky flag
      rating 2 (unsure)    → 1 pt  → conf_ratio 0.5 → weight ~1.25 at 100% acc
      rating 3 (kinda sure)→ 2 pts → conf_ratio 1.0 → weight ~0.50 at 100% acc
      rating 4 (sure)      → 3 pts → conf_ratio 1.5 → weight ~0.30 at 100% acc
      no rating / exam mode→ 2 pts → same as kinda sure
    """
    lucky      = q.get("lucky", 0)
    conf_sum   = q.get("conf_sum", 0)
    wrong      = q["wrong"] + lucky
    real_right = max(0, q["right"] - lucky)
    total      = wrong + real_right
    if total == 0:
        return 2.0

    base_accuracy = real_right / total

    # conf_ratio: how confident are the correct answers on average?
    # 1.0 = kinda sure (baseline), <1 = less confident, >1 = more confident
    if real_right > 0 and conf_sum > 0:
        conf_ratio = conf_sum / (real_right * 2)
    else:
        conf_ratio = 1.0  # backward-compatible: unrated answers treated as kinda sure

    effective_accuracy = base_accuracy * conf_ratio
    return max(0.3, 2.0 - effective_accuracy * 1.5)


def weighted_sample(pool: list, k: int) -> list:
    """Sample k items from pool using adaptive weights (without replacement)."""
    if k >= len(pool):
        result = pool[:]
        random.shuffle(result)
        return result
    weights = [question_weight(q) for q in pool]
    selected = []
    remaining = list(zip(pool, weights))
    for _ in range(k):
        if not remaining:
            break
        items, wts = zip(*remaining)
        chosen = random.choices(items, weights=wts, k=1)[0]
        selected.append(chosen)
        remaining = [(it, w) for it, w in remaining if it is not chosen]
    return selected

# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

def show_dashboard(all_questions: list) -> None:
    """Print a progress table grouped by subsection."""
    ordered_subs = []
    for sk in sorted(SECTION_MAP.keys()):
        ordered_subs.extend(SECTION_MAP[sk]["subsections"])

    by_sub = {s: [] for s in ordered_subs}
    for q in all_questions:
        if q["section"] in by_sub:
            by_sub[q["section"]].append(q)

    print(f"\n{DIVIDER}")
    print("  YOUR PROGRESS")
    print(DIVIDER)
    print(f"  {'Section':<8} {'Topic':<38} {'Q':>4} {'Right':>5} {'Wrong':>5} {'Lucky':>5} {'Conf':>5} {'%':>6}")
    print(THIN_DIV)

    for sk in sorted(SECTION_MAP.keys()):
        for sub in SECTION_MAP[sk]["subsections"]:
            qs       = by_sub.get(sub, [])
            total    = len(qs)
            right    = sum(q["right"] for q in qs)
            wrong    = sum(q["wrong"] for q in qs)
            lucky    = sum(q.get("lucky", 0) for q in qs)
            conf_sum = sum(q.get("conf_sum", 0) for q in qs)
            asked    = right + wrong
            pct      = f"{right/asked*100:.0f}%" if asked > 0 else "—"
            real_r   = max(0, right - lucky)
            conf_disp = f"{conf_sum/real_r+1:.1f}" if (real_r > 0 and conf_sum > 0) else ("—" if real_r == 0 else "3.0")
            name     = SECTION_MAP[sk]["sub_names"].get(sub, "")
            print(f"  {sub:<8} {name:<38} {total:>4} {right:>5} {wrong:>5} {lucky:>5} {conf_disp:>5} {pct:>6}")

    print(THIN_DIV)
    total_q  = len(all_questions)
    total_r  = sum(q["right"] for q in all_questions)
    total_w  = sum(q["wrong"] for q in all_questions)
    total_l  = sum(q.get("lucky", 0) for q in all_questions)
    total_cs = sum(q.get("conf_sum", 0) for q in all_questions)
    asked    = total_r + total_w
    pct_all  = f"{total_r/asked*100:.0f}%" if asked > 0 else "—"
    real_r_all = max(0, total_r - total_l)
    conf_all   = f"{total_cs/real_r_all+1:.1f}" if (real_r_all > 0 and total_cs > 0) else ("—" if real_r_all == 0 else "3.0")
    print(f"  {'TOTAL':<8} {'':<38} {total_q:>4} {total_r:>5} {total_w:>5} {total_l:>5} {conf_all:>5} {pct_all:>6}")
    print(DIVIDER)

# ---------------------------------------------------------------------------
# Menus
# ---------------------------------------------------------------------------

def _input(prompt: str) -> str:
    """Wrapper around input() that prints help on H/? and returns stripped upper."""
    while True:
        try:
            val = input(prompt).strip()
        except EOFError:
            return "Q"
        if val.upper() in ("H", "?"):
            print(HELP_TEXT)
        else:
            return val.upper()


def reset_data() -> None:
    """Zero out all tracking columns in Excel and delete the session log."""
    print(f"\n{DIVIDER}")
    print("  RESET DATA")
    print(DIVIDER)
    print("  This will clear ALL right/wrong/times-asked stats in the")
    print("  question bank and delete session_log.csv.")
    print()
    confirm = input('  Type  RESET DATA  to confirm, or press Enter to cancel: ').strip()
    if confirm != "RESET DATA":
        print("  Cancelled.")
        return

    print("  Resetting Excel stats...", flush=True)
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value is None:
                continue
            row[7].value = 0   # # Of times asked
            row[8].value = 0   # RIGHT
            row[9].value = 0   # WRONG
            if len(row) > 11:
                row[11].value = 0  # LUCKY
            if len(row) > 12:
                row[12].value = 0  # CONF_SUM
        wb.save(EXCEL_PATH)
        wb.close()
        print("  Excel stats cleared.")
    except PermissionError:
        print("  ERROR: Excel file is open. Close it and try again.")
        return
    except Exception as exc:
        print(f"  ERROR resetting Excel: {exc}")
        return

    if os.path.exists(LOG_PATH):
        try:
            os.remove(LOG_PATH)
            print("  Session log deleted.")
        except Exception as exc:
            print(f"  ERROR deleting log: {exc}")
    else:
        print("  No session log found (nothing to delete).")

    print(f"\n  All data reset. Starting fresh!")
    print(DIVIDER)


def choose_mode() -> str:
    """Return mode code: F, S, W, B, V, or R."""
    while True:
        print(f"\n{DIVIDER}")
        print("  SELECT MODE")
        print(DIVIDER)
        print("  F  Full Exam Mode  (75 q, real SIE distribution)")
        print("  S  Study Mode      (choose sections + target count)")
        print("  W  Weak Areas      (questions you've gotten wrong)")
        print("  B  Bookmarks       (questions you've bookmarked)")
        print("  V  View History    (past session scores)")
        print("  R  Reset Data      (clear all stats and logs)")
        print("  H  Help            Q  Quit")
        print(THIN_DIV)
        choice = _input("  Your choice: ")
        if choice in ("F", "S", "W", "B", "V", "R"):
            return choice
        if choice == "Q":
            raise SystemExit
        if choice not in ("H", "?"):
            print("  Invalid — please enter F, S, W, B, V, or R.")


def choose_main_section() -> str:
    """Return a section key '1'–'4' or 'ALL'."""
    while True:
        print(f"\n{DIVIDER}")
        print("  SELECT SECTION")
        print(DIVIDER)
        for k, v in SECTION_MAP.items():
            print(f"  {k}  Section {k}: {v['name']}")
        print("  A  All Sections")
        print("  H  Help            Q  Back")
        print(THIN_DIV)
        choice = _input("  Your choice: ")
        if choice in SECTION_MAP:
            return choice
        if choice == "A":
            return "ALL"
        if choice == "Q":
            return "Q"
        if choice not in ("H", "?"):
            print("  Invalid — enter a section number (1–4), A, or Q.")


def checklist_menu(items: list, counts: dict, title: str) -> list:
    """
    Interactive toggle checklist. Returns list of selected item strings.
    items  — list of strings (subsection codes)
    counts — dict mapping item -> int (question count for display)
    title  — header text
    """
    selected = set()

    while True:
        print(f"\n{DIVIDER}")
        print(f"  {title}")
        print(DIVIDER)
        for i, item in enumerate(items, 1):
            mark  = "X" if item in selected else " "
            count = counts.get(item, 0)
            name  = ""
            for sk, sv in SECTION_MAP.items():
                if item in sv["sub_names"]:
                    name = sv["sub_names"][item]
                    break
            print(f"  [{mark}] {i}.  {item}  {name:<40}  ({count} q)")
        print(THIN_DIV)
        print("  A  Select All    C  Clear All    Enter  Confirm")
        print("  H  Help          Q  Back")
        print(THIN_DIV)

        raw = input("  Toggle (number/A/C/Enter/Q): ").strip().upper()

        if raw == "":
            if not selected:
                print("  Select at least one subsection first.")
            else:
                return list(selected)
        elif raw == "A":
            selected = set(items)
        elif raw == "C":
            selected = set()
        elif raw == "Q":
            return []
        elif raw in ("H", "?"):
            print(HELP_TEXT)
        else:
            try:
                idx = int(raw) - 1
                if 0 <= idx < len(items):
                    item = items[idx]
                    if item in selected:
                        selected.discard(item)
                    else:
                        selected.add(item)
                else:
                    print(f"  Invalid — enter a number between 1 and {len(items)}.")
            except ValueError:
                print("  Invalid input.")


def get_target_count(available: int, adaptive_on: bool) -> tuple:
    """
    Ask user how many questions and whether adaptive mode is on.
    Returns (target: int, adaptive: bool).
    """
    print(f"\n  Questions available: {available}")
    print(f"  Adaptive learning:  {'ON' if adaptive_on else 'OFF'}  (toggle with A)")

    while True:
        raw = input(f"  How many questions? (Enter = all {available}, A = toggle adaptive): ").strip().upper()
        if raw == "":
            return available, adaptive_on
        if raw == "A":
            adaptive_on = not adaptive_on
            print(f"  Adaptive learning: {'ON' if adaptive_on else 'OFF'}")
            continue
        if raw in ("H", "?"):
            print(HELP_TEXT)
            continue
        try:
            n = int(raw)
            if n <= 0:
                print("  Enter a positive number.")
            else:
                return n, adaptive_on
        except ValueError:
            print("  Invalid — enter a number or press Enter.")

# ---------------------------------------------------------------------------
# Question gathering
# ---------------------------------------------------------------------------

def filter_by_subsections(all_questions: list, subsections: list) -> list:
    return [q for q in all_questions if q["section"] in subsections]


def gather_study_questions(all_questions: list, selected_subs: list,
                           section_key: str, target: int,
                           adaptive: bool) -> list:
    """
    Build a question list for Study Mode.
    Fills from neighboring subsections (same main section) if needed.
    """
    pool = filter_by_subsections(all_questions, selected_subs)

    if len(pool) < target and section_key != "ALL":
        all_subs  = SECTION_MAP[section_key]["subsections"]
        neighbors = [s for s in all_subs if s not in selected_subs]
        for neighbor in neighbors:
            if len(pool) >= target:
                break
            pool += filter_by_subsections(all_questions, [neighbor])
        if len(pool) < target:
            print(f"\n  [Note] Only {len(pool)} questions available in Section {section_key} — using all of them.")

    if adaptive:
        return weighted_sample(pool, target)
    else:
        sampled = pool[:] if target >= len(pool) else random.sample(pool, target)
        random.shuffle(sampled)
        return sampled


def gather_exam_questions(all_questions: list, adaptive: bool) -> list:
    """Build a 75-question list matching the real SIE exam distribution."""
    result = []
    for sk, count in EXAM_DISTRIBUTION.items():
        section_subs = SECTION_MAP[sk]["subsections"]
        pool = filter_by_subsections(all_questions, section_subs)
        if len(pool) == 0:
            print(f"  [Warning] Section {sk} has no questions — skipping its {count} slots.")
            continue
        if len(pool) < count:
            print(f"  [Note] Section {sk} only has {len(pool)} questions (need {count}) — using all available.")
        if adaptive:
            sampled = weighted_sample(pool, count)
        else:
            sampled = pool[:] if count >= len(pool) else random.sample(pool, count)
        result.extend(sampled)

    random.shuffle(result)
    return result


def missed_to_questions(missed_questions: list) -> list:
    """Convert missed_questions entries back to question dicts for re-quizzing."""
    return [{
        "row":         mq["row"],
        "section":     mq["section"],
        "question":    mq["question"],
        "choices":     mq["choices"],
        "correct":     mq["correct"],
        "explanation": mq.get("explanation", ""),
        "times_asked": 0,
        "right":       0,
        "wrong":       0,
    } for mq in missed_questions]

# ---------------------------------------------------------------------------
# Quiz engine
# ---------------------------------------------------------------------------

def _copy_to_clipboard(q: dict) -> None:
    """Copy question + all answer choices (A–D order) + correct marker to clipboard."""
    sec_key  = q["section"].split(".")[0]
    sub_name = SECTION_MAP.get(sec_key, {}).get("sub_names", {}).get(q["section"], q["section"])
    lines = [f"[SIE Section {q['section']} — {sub_name}]", "", f"Q: {q['question']}", ""]
    for letter in ["A", "B", "C", "D"]:
        marker = "  <- CORRECT" if letter == q["correct"] else ""
        lines.append(f"  {letter}. {q['choices'][letter]}{marker}")
    if q.get("explanation"):
        lines += ["", f"  Explanation: {q['explanation']}"]
    text = "\n".join(lines)
    try:
        subprocess.run("clip", input=text.encode("utf-8"), check=True, shell=True)
        print("  Copied to clipboard! Paste into Claude to get help.")
    except Exception:
        print("  (Could not copy to clipboard — try again.)")


def _fmt_elapsed(seconds: float) -> str:
    """Format elapsed seconds as Xm Ys."""
    m = int(seconds) // 60
    s = int(seconds) % 60
    return f"{m}m {s:02d}s" if m > 0 else f"{s}s"


def _wrap(text: str, width: int = 60, indent: str = "  ") -> list:
    """Word-wrap text to width, returning list of indented lines."""
    words = text.split()
    line, lines = "", []
    for w in words:
        if len(line) + len(w) + 1 > width:
            lines.append(indent + line)
            line = w
        else:
            line = (line + " " + w).strip()
    if line:
        lines.append(indent + line)
    return lines


def run_quiz(questions: list, bookmarks: set = None, exam_mode: bool = False) -> tuple:
    """
    Run the quiz loop.
    Returns (correct_count, total, section_results, missed_questions, elapsed_seconds).

    missed_questions entries include wrong answers, IDK, and lucky guesses
    (correct answers where confidence == 1).
    """
    correct_count    = 0
    total            = len(questions)
    section_results  = {}
    missed_questions = []
    session_start    = time.time()
    last_answered    = 0

    for i, q in enumerate(questions, 1):
        print(f"\n{DIVIDER}")
        bm_tag = "  [BOOKMARKED]" if (bookmarks is not None and q["row"] in bookmarks) else ""
        print(f"  Question {i} of {total}   |   Section {q['section']}{bm_tag}")
        print(DIVIDER)

        for ln in _wrap(q["question"]):
            print(ln)
        print()

        # Shuffle answer choices
        letters = ["A", "B", "C", "D"]
        random.shuffle(letters)
        mapping = {str(num): letter for num, letter in enumerate(letters, 1)}

        for num, letter in mapping.items():
            print(f"    {num}. {q['choices'][letter]}")
        print()

        correct_letter = q["correct"]
        correct_text   = q["choices"][correct_letter]

        # --- Get answer ---
        answer   = None
        was_idk  = False
        end_exam = False
        if exam_mode:
            ans_prompt = "  Your answer (1–4, IDK=skip, B=bookmark, C=copy, END TEST=end, H=help, Q=quit): "
        else:
            ans_prompt = "  Your answer (1–4, IDK=skip, C=copy, H=help, Q=quit): "
        while answer is None:
            raw = _input(ans_prompt)
            if raw == "Q":
                print("\n  Returning to main menu...")
                return correct_count, last_answered, section_results, missed_questions, time.time() - session_start
            if exam_mode and raw in ("END", "END TEST"):
                end_exam = True
                break
            if raw == "IDK":
                answer  = "IDK"
                was_idk = True
            elif raw == "B" and exam_mode:
                if bookmarks is not None:
                    if q["row"] in bookmarks:
                        bookmarks.discard(q["row"])
                        save_bookmarks(bookmarks)
                        print("  Bookmark removed.")
                    else:
                        bookmarks.add(q["row"])
                        save_bookmarks(bookmarks)
                        print("  Bookmarked!")
                else:
                    print("  (Bookmarks not available.)")
            elif raw == "C":
                _copy_to_clipboard(q)
            elif raw in ("1", "2", "3", "4"):
                answer = raw
            elif raw not in ("H", "?"):
                if exam_mode:
                    print("  Invalid — enter 1–4, IDK, B, C, END TEST, H, or Q.")
                else:
                    print("  Invalid — enter 1–4, IDK, C, H, or Q.")

        if end_exam:
            break

        # --- Evaluate answer ---
        if was_idk:
            is_correct = False
            if not exam_mode:
                print(f"\n  Skipped — correct answer:  {correct_text}")
        else:
            chosen_letter = mapping[answer]
            is_correct    = (chosen_letter == correct_letter)
            if is_correct:
                correct_count += 1
                if not exam_mode:
                    print(f"\n  CORRECT!  {correct_text}")
            else:
                if not exam_mode:
                    chosen_text = q["choices"][chosen_letter]
                    print(f"\n  INCORRECT.  You chose:     {chosen_text}")
                    print(f"  Correct answer:  {correct_text}")

        # Show explanation if available (wrong / IDK only)
        if not exam_mode and not is_correct and q.get("explanation"):
            print(f"\n  WHY: {q['explanation']}")

        # --- Confidence + Bookmark (correct answers only, not in exam mode) ---
        confidence = None
        if is_correct and not exam_mode:
            print("  Confidence?  1 guessed · 2 unsure · 3 kinda sure · 4 sure · Enter skip")
            print("               Add ,B to bookmark at same time  (e.g. 3,B  or  B,2)")
            while True:
                conf_raw = input("  > ").strip().upper()
                tokens = [t.strip() for t in conf_raw.split(",") if t.strip()]
                if not all(t in ("1", "2", "3", "4", "B") for t in tokens):
                    print("  Invalid — use 1–4, B, or combine like 3,B")
                    continue
                rating_tok = next((t for t in tokens if t in ("1", "2", "3", "4")), None)
                if "B" in tokens:
                    if bookmarks is not None:
                        if q["row"] in bookmarks:
                            bookmarks.discard(q["row"])
                            save_bookmarks(bookmarks)
                            print("  Bookmark removed.")
                        else:
                            bookmarks.add(q["row"])
                            save_bookmarks(bookmarks)
                            print("  Bookmarked!")
                    else:
                        print("  (Bookmarks not available.)")
                confidence = int(rating_tok) if rating_tok else None
                break

        # Track missed: wrong, IDK, or correct-but-guessed (lucky guess)
        was_lucky = is_correct and confidence == 1
        if not is_correct or was_lucky:
            missed_questions.append({
                "row":          q["row"],
                "section":      q["section"],
                "question":     q["question"],
                "choices":      q["choices"],
                "correct":      correct_letter,
                "correct_text": correct_text,
                "explanation":  q.get("explanation", ""),
                "was_idk":      was_idk,
                "was_lucky":    was_lucky,
                "confidence":   confidence,
            })

        # Track per-section
        sec_key = q["section"].split(".")[0]
        if sec_key not in section_results:
            section_results[sec_key] = [0, 0]
        section_results[sec_key][1] += 1
        if is_correct:
            section_results[sec_key][0] += 1

        # Save to Excel
        conf_pts = {1: 0, 2: 1, 3: 2, 4: 3}.get(confidence, 2) if (is_correct and not was_idk) else 0
        save_result(q["row"], is_correct, lucky=was_lucky, conf_pts=conf_pts)
        last_answered = i
        print(THIN_DIV)

    return correct_count, last_answered, section_results, missed_questions, time.time() - session_start


def show_score(correct: int, total: int,
               section_results: dict = None, exam_mode: bool = False,
               missed_questions: list = None, elapsed: float = None) -> None:
    """Print final score, timing, and WHAT TO STUDY block."""
    print(f"\n{DIVIDER}")
    print("  RESULTS")
    print(DIVIDER)

    if total == 0:
        print("  No questions were answered.")
        print(DIVIDER)
        return

    pct = correct / total * 100

    # Timing
    if elapsed is not None and elapsed > 0:
        avg_per_q = elapsed / total
        print(f"  Time: {_fmt_elapsed(elapsed)}   (avg {avg_per_q:.1f}s / question)")
        print(THIN_DIV)

    if exam_mode and section_results:
        for sk in sorted(SECTION_MAP.keys()):
            if sk in section_results:
                sc, st = section_results[sk]
                sp = sc / st * 100 if st > 0 else 0
                name = SECTION_MAP[sk]["name"]
                print(f"  Section {sk}: {sc:>2}/{st:<2}  ({sp:.1f}%)  {name}")
        print(THIN_DIV)

    print(f"  Overall:  {correct} / {total}  ({pct:.1f}%)")

    if exam_mode:
        result = "PASS" if pct >= PASS_THRESHOLD * 100 else "FAIL"
        print(f"\n  SIE Pass threshold: {PASS_THRESHOLD*100:.0f}%   Result: {result}")

    print(DIVIDER)

    # WHAT TO STUDY block
    if missed_questions:
        print(f"\n{DIVIDER}")
        print(f"  WHAT TO STUDY  ({len(missed_questions)} question(s) to review)")
        print(DIVIDER)

        by_section: dict = {}
        for mq in missed_questions:
            by_section.setdefault(mq["section"], []).append(mq)

        for sec in sorted(by_section.keys()):
            sec_key  = sec.split(".")[0]
            sub_name = SECTION_MAP.get(sec_key, {}).get("sub_names", {}).get(sec, sec)
            print(f"\n  [Section {sec} — {sub_name}]")
            for mq in by_section[sec]:
                if mq.get("was_lucky"):
                    tag = "  (Lucky Guess)"
                elif mq["was_idk"]:
                    tag = "  (IDK)"
                else:
                    tag = "  (Wrong)"
                for ln in _wrap(mq["question"], width=56):
                    prefix = f"  Q{tag}:" if ln == _wrap(mq["question"], width=56)[0] else "         "
                    print(f"{prefix} {ln.strip()}")
                print(f"  A: {mq['correct_text']}")
                if mq.get("explanation"):
                    print(f"  WHY: {mq['explanation']}")
                print(THIN_DIV)

        print()


def show_history() -> None:
    """Display past session scores from session_log.csv."""
    print(f"\n{DIVIDER}")
    print("  SESSION HISTORY")
    print(DIVIDER)

    if not os.path.exists(LOG_PATH):
        print("  No session history yet — complete a quiz to start tracking.")
        print(DIVIDER)
        return

    try:
        with open(LOG_PATH, "r", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
    except Exception as exc:
        print(f"  Could not read session log: {exc}")
        print(DIVIDER)
        return

    if not rows:
        print("  No sessions recorded yet.")
        print(DIVIDER)
        return

    recent = rows[-20:]
    print(f"  {'#':<4} {'Date':<12} {'Mode':<12} {'Score':<10} {'%':>6}")
    print(THIN_DIV)
    for i, r in enumerate(recent, 1):
        try:
            score = f"{r['correct']}/{r['total']}"
            pct   = f"{float(r['pct']):.1f}%"
            mode  = r.get("mode", "?")[:11]
            date  = r.get("date", "?")
            print(f"  {i:<4} {date:<12} {mode:<12} {score:<10} {pct:>6}")
        except (KeyError, ValueError):
            continue

    if len(rows) >= 2:
        last5 = rows[-5:]
        try:
            avg = sum(float(r["pct"]) for r in last5) / len(last5)
            print(THIN_DIV)
            print(f"  Last {len(last5)}-session avg: {avg:.1f}%")
        except (KeyError, ValueError):
            pass

    print(DIVIDER)


def exam_review(questions_shown: list, missed_questions: list, bookmarks: set) -> None:
    """After a Full Exam: let the user browse bookmarked and IDK questions with answers revealed."""
    idk_rows = {mq["row"] for mq in missed_questions if mq.get("was_idk")}
    bm_rows  = {q["row"] for q in questions_shown if q["row"] in (bookmarks or set())}

    review_list = []
    for q in questions_shown:
        is_bm  = q["row"] in bm_rows
        is_idk = q["row"] in idk_rows
        if is_bm or is_idk:
            review_list.append((is_bm, is_idk, q))

    if not review_list:
        return

    bm_count  = sum(1 for bm, _,  __ in review_list if bm)
    idk_count = sum(1 for  _, idk, __ in review_list if idk)

    print(f"\n{DIVIDER}")
    print("  EXAM REVIEW — BOOKMARKED & IDK QUESTIONS")
    print(DIVIDER)
    if bm_count:
        print(f"  [B]  Bookmarked:    {bm_count} question(s)")
    if idk_count:
        print(f"  [I]  IDK (skipped): {idk_count} question(s)")
    print()

    for i, (is_bm, is_idk, q) in enumerate(review_list, 1):
        tag     = "[BI]" if (is_bm and is_idk) else "[B] " if is_bm else "[I] "
        preview = (q["question"][:52] + "...") if len(q["question"]) > 52 else q["question"]
        print(f"  {tag} {i:2}.  Sec {q['section']}  {preview}")

    print(THIN_DIV)
    raw = input("  Review which? (e.g. 1,3  or  ALL  or Enter to skip): ").strip().upper()

    if not raw:
        return

    if raw == "ALL":
        indices = list(range(len(review_list)))
    else:
        indices = []
        for part in raw.split(","):
            try:
                idx = int(part.strip()) - 1
                if 0 <= idx < len(review_list):
                    indices.append(idx)
            except ValueError:
                pass
        if not indices:
            print("  No valid question numbers entered — skipping review.")
            return

    review_correct = 0
    for n, idx in enumerate(indices, 1):
        is_bm, is_idk, q = review_list[idx]
        tag_label = ("BOOKMARKED" + (" | IDK" if is_idk else "")) if is_bm else "IDK"

        print(f"\n{DIVIDER}")
        print(f"  Review {n}/{len(indices)}   Section {q['section']}   {tag_label}")
        print(DIVIDER)

        for ln in _wrap(q["question"]):
            print(ln)
        print()

        # Shuffle choices just like the real exam
        letters = ["A", "B", "C", "D"]
        random.shuffle(letters)
        mapping = {str(num): letter for num, letter in enumerate(letters, 1)}

        for num, letter in mapping.items():
            print(f"    {num}. {q['choices'][letter]}")
        print()

        correct_letter = q["correct"]
        correct_text   = q["choices"][correct_letter]

        answer = None
        while answer is None:
            raw = input("  Your answer (1–4, Enter=reveal): ").strip().upper()
            if raw in ("1", "2", "3", "4"):
                answer = raw
            elif raw == "":
                answer = "REVEAL"
            else:
                print("  Invalid — enter 1–4 or press Enter to reveal.")

        if answer == "REVEAL":
            print(f"\n  Answer: {correct_text}")
            save_result(q["row"], correct=False, conf_pts=0)
        else:
            chosen_letter = mapping[answer]
            is_correct    = (chosen_letter == correct_letter)
            if is_correct:
                review_correct += 1
                print(f"\n  CORRECT!  {correct_text}")
                save_result(q["row"], correct=True, conf_pts=2)
            else:
                print(f"\n  INCORRECT.  You chose: {q['choices'][chosen_letter]}")
                print(f"  Correct answer:        {correct_text}")
                save_result(q["row"], correct=False, conf_pts=0)

        if q.get("explanation"):
            print(f"\n  WHY: {q['explanation']}")

        print(THIN_DIV)

    print(f"\n  Review complete.  {review_correct}/{len(indices)} correct on re-attempt.")
    print(DIVIDER)

# ---------------------------------------------------------------------------
# Session log helpers
# ---------------------------------------------------------------------------

def _sections_label(mode: str, selected_subs: list) -> str:
    if mode == "F":
        return "Full Exam"
    if mode == "W":
        return "Weak Areas"
    return "+".join(sorted(selected_subs))

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    # Fix Windows terminal Unicode issues
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    print(f"\n{DIVIDER}")
    print("  SIE EXAM QUIZZER")
    print("  Type H at any prompt for help.")
    print(DIVIDER)

    print("  Loading question bank...", flush=True)
    all_questions = load_all_questions()
    if not all_questions:
        print("\n  No questions found in the Excel file. Add questions and re-run.")
        sys.exit(0)

    bookmarks = load_bookmarks()
    show_dashboard(all_questions)

    adaptive_default = True

    try:
        while True:
            mode = choose_mode()
            missed = []

            # --- View History ---
            if mode == "V":
                show_history()
                continue

            # --- Reset ---
            if mode == "R":
                reset_data()
                all_questions = load_all_questions()
                show_dashboard(all_questions)
                continue

            # --- Full Exam Mode ---
            if mode == "F":
                print(f"\n  Building Full Exam (75 questions)...")
                questions = gather_exam_questions(all_questions, adaptive=True)
                if not questions:
                    print("  No questions available for Full Exam Mode. Add questions and re-run.")
                    continue
                print(f"  {len(questions)} questions loaded. Good luck!\n")
                correct, total_q, section_results, missed, elapsed = run_quiz(questions, bookmarks, exam_mode=True)
                show_score(correct, total_q, section_results, exam_mode=True,
                           missed_questions=missed, elapsed=elapsed)
                append_session_log("Full Exam", "All", 75, correct, total_q)
                exam_review(questions[:total_q], missed, bookmarks)

            # --- Bookmarks Mode ---
            elif mode == "B":
                bq = [q for q in all_questions if q["row"] in bookmarks]
                if not bq:
                    print("\n  No bookmarks yet — press B during a quiz to bookmark a question.")
                    continue
                print(f"\n  {len(bq)} bookmarked question(s) loaded.")
                random.shuffle(bq)
                correct, total_q, section_results, missed, elapsed = run_quiz(bq, bookmarks)
                show_score(correct, total_q, missed_questions=missed, elapsed=elapsed)
                # Offer to remove correctly-answered bookmarks
                missed_rows    = {mq["row"] for mq in missed}
                got_right_rows = {q["row"] for q in bq[:total_q]} - missed_rows
                if got_right_rows:
                    clear = _input(f"  Remove {len(got_right_rows)} correctly-answered bookmark(s)? (Y/N): ")
                    if clear == "Y":
                        bookmarks -= got_right_rows
                        save_bookmarks(bookmarks)
                        print(f"  {len(got_right_rows)} bookmark(s) removed.")
                append_session_log("Bookmarks", "Bookmarks", len(bq), correct, total_q)

            # --- Weak Areas Mode ---
            elif mode == "W":
                weak = [q for q in all_questions if q["wrong"] > 0]
                if not weak:
                    print("\n  No wrong answers yet — keep studying!")
                    continue
                print(f"\n  {len(weak)} weak-area questions found.")
                target, _ = get_target_count(len(weak), adaptive_default)
                sampled = random.sample(weak, min(target, len(weak)))
                random.shuffle(sampled)
                correct, total_q, section_results, missed, elapsed = run_quiz(sampled, bookmarks)
                show_score(correct, total_q, missed_questions=missed, elapsed=elapsed)
                append_session_log("Weak Areas", "Weak", target, correct, total_q)

            # --- Study Mode ---
            else:
                sec_key = choose_main_section()
                if sec_key == "Q":
                    continue

                if sec_key == "ALL":
                    all_subs = []
                    for sv in SECTION_MAP.values():
                        all_subs.extend(sv["subsections"])
                    counts = {sub: sum(1 for q in all_questions if q["section"] == sub)
                              for sub in all_subs}
                    selected_subs = checklist_menu(all_subs, counts, "SELECT SUBSECTIONS (All Sections)")
                    effective_section = "ALL"
                else:
                    subs   = SECTION_MAP[sec_key]["subsections"]
                    counts = {sub: sum(1 for q in all_questions if q["section"] == sub)
                              for sub in subs}
                    selected_subs = checklist_menu(subs, counts,
                                                   f"SELECT SUBSECTIONS — Section {sec_key}")
                    effective_section = sec_key

                if not selected_subs:
                    continue

                pool_size = len(filter_by_subsections(all_questions, selected_subs))
                if effective_section != "ALL":
                    all_sec_subs     = SECTION_MAP[effective_section]["subsections"]
                    total_in_section = sum(1 for q in all_questions
                                          if q["section"] in all_sec_subs)
                    if total_in_section > pool_size:
                        print(f"\n  ({total_in_section - pool_size} more questions available from "
                              f"other subsections in Section {effective_section} for auto-fill)")

                target, adaptive_default = get_target_count(pool_size, adaptive_default)
                questions = gather_study_questions(
                    all_questions, selected_subs, effective_section,
                    target, adaptive_default
                )
                if not questions:
                    print("  No questions found for that selection. Try a different section.")
                    continue

                correct, total_q, section_results, missed, elapsed = run_quiz(questions, bookmarks)
                show_score(correct, total_q, missed_questions=missed, elapsed=elapsed)
                append_session_log(
                    "Study", _sections_label(mode, selected_subs),
                    target, correct, total_q
                )

            # --- Re-quiz missed questions ---
            if missed:
                rq = _input(f"  Re-quiz {len(missed)} missed question(s)? (Y/N): ")
                if rq == "Y":
                    rq_questions = missed_to_questions(missed)
                    random.shuffle(rq_questions)
                    rq_correct, rq_total, rq_sec, rq_missed, rq_elapsed = run_quiz(rq_questions, bookmarks)
                    show_score(rq_correct, rq_total, missed_questions=rq_missed, elapsed=rq_elapsed)

            # Continue prompt
            print()
            again = _input("  Study another session? (Y / N): ")
            if again != "Y":
                break

            # Reload so updated stats are reflected in next session
            all_questions = load_all_questions()
            show_dashboard(all_questions)

    except (KeyboardInterrupt, SystemExit):
        pass

    print("\n  Goodbye! Good luck on the SIE.\n")


if __name__ == "__main__":
    main()
