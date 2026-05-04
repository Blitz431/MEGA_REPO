"""
SIE Exam Quizzer — Web App
Run:  python web_app.py
Then open http://localhost:5000 in your browser.
"""

import os
import json
import random
import datetime
import csv
import threading
from flask import Flask, jsonify, request, render_template

try:
    import openpyxl
except ImportError:
    raise SystemExit("ERROR: openpyxl is not installed. Run:  pip install openpyxl")

BASE           = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH     = os.path.join(BASE, "QUESTION BANK AND TRACKER.xlsx")
LOG_PATH       = os.path.join(BASE, "session_log.csv")
BOOKMARKS_PATH = os.path.join(BASE, "bookmarks.json")

SECTION_MAP = {
    "1": {
        "name": "Knowledge of Capital Markets",
        "subsections": ["1.1", "1.2", "1.3"],
        "sub_names": {
            "1.1": "Regulatory Entities, Agencies & Market Participants",
            "1.2": "Market Structure",
            "1.3": "Economic Factors",
        },
    },
    "2": {
        "name": "Understanding Products and Their Risks",
        "subsections": ["2.1", "2.2", "2.3", "2.4", "2.5"],
        "sub_names": {
            "2.1": "Equity Securities",
            "2.2": "Debt Securities",
            "2.3": "Packaged Products",
            "2.4": "Options",
            "2.5": "Alternative Investments & Other Products",
        },
    },
    "3": {
        "name": "Understanding Trading, Customer Accounts & Prohibited Activities",
        "subsections": ["3.1", "3.2", "3.3"],
        "sub_names": {
            "3.1": "Trading Securities",
            "3.2": "Customer Accounts & Compliance",
            "3.3": "Prohibited Activities",
        },
    },
    "4": {
        "name": "Overview of the Regulatory Framework",
        "subsections": ["4.1", "4.2", "4.3"],
        "sub_names": {
            "4.1": "SROs and Regulatory Bodies",
            "4.2": "Registration & Licensing",
            "4.3": "Key Regulations & Rules",
        },
    },
}

EXAM_DISTRIBUTION = {"1": 12, "2": 33, "3": 23, "4": 7}

_excel_lock = threading.Lock()

app = Flask(__name__)


# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------

def load_all_questions():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    questions = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 10:
            continue
        section_raw, question, a, b, c, d, correct, times, right, wrong = (row[i] for i in range(10))
        explanation = str(row[10]).strip() if len(row) > 10 and row[10] is not None else ""
        if section_raw is None:
            continue
        try:
            section_str = f"{float(section_raw):.1f}"
        except (ValueError, TypeError):
            section_str = str(section_raw).strip()
        if any(v is None for v in [question, a, b, c, d, correct]):
            continue
        correct = str(correct).strip().upper()
        if correct not in {"A", "B", "C", "D"}:
            continue
        questions.append({
            "row":         row_idx,
            "section":     section_str,
            "question":    str(question),
            "choices":     {"A": str(a), "B": str(b), "C": str(c), "D": str(d)},
            "correct":     correct,
            "explanation": explanation,
            "times_asked": int(times) if times is not None else 0,
            "right":       int(right) if right is not None else 0,
            "wrong":       int(wrong) if wrong is not None else 0,
        })
    wb.close()
    return questions


def save_result(row, correct):
    with _excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active
            def inc(col):
                v = ws.cell(row=row, column=col).value
                ws.cell(row=row, column=col).value = (int(v) if v is not None else 0) + 1
            inc(8)
            inc(9) if correct else inc(10)
            wb.save(EXCEL_PATH)
            wb.close()
            return True
        except Exception:
            return False


def load_bookmarks():
    if not os.path.exists(BOOKMARKS_PATH):
        return set()
    try:
        with open(BOOKMARKS_PATH, "r", encoding="utf-8") as f:
            return set(json.load(f))
    except Exception:
        return set()


def save_bookmarks(bookmarks):
    with open(BOOKMARKS_PATH, "w", encoding="utf-8") as f:
        json.dump(sorted(bookmarks), f)


def question_weight(q):
    total = q["wrong"] + q["right"]
    if total == 0:
        return 2.0
    return max(0.5, 2.0 - (q["right"] / total) * 1.5)


def weighted_sample(pool, k):
    if k >= len(pool):
        result = pool[:]
        random.shuffle(result)
        return result
    weights = [question_weight(q) for q in pool]
    selected = []
    remaining = list(zip(pool, weights))
    for _ in range(k):
        if not remaining:
            break
        items, wts = zip(*remaining)
        chosen = random.choices(items, weights=wts, k=1)[0]
        selected.append(chosen)
        remaining = [(it, w) for it, w in remaining if it is not chosen]
    return selected


def append_session_log(mode, sections, target, correct, total):
    pct = (correct / total * 100) if total > 0 else 0.0
    write_header = not os.path.exists(LOG_PATH)
    try:
        with open(LOG_PATH, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if write_header:
                writer.writerow(["date", "mode", "sections", "target", "correct", "total", "pct"])
            writer.writerow([
                datetime.date.today().isoformat(),
                mode, sections, target, correct, total, f"{pct:.1f}",
            ])
    except Exception:
        pass


def format_question(q, bookmarks):
    letters = ["A", "B", "C", "D"]
    random.shuffle(letters)
    return {
        "row":        q["row"],
        "section":    q["section"],
        "question":   q["question"],
        "choices":    [{"key": letter, "text": q["choices"][letter]} for letter in letters],
        "correct":    q["correct"],
        "explanation": q["explanation"],
        "bookmarked": q["row"] in bookmarks,
    }


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/section_map")
def section_map_route():
    return jsonify(SECTION_MAP)


@app.route("/api/dashboard")
def dashboard():
    qs        = load_all_questions()
    bookmarks = load_bookmarks()
    sections  = []
    for sk in sorted(SECTION_MAP.keys()):
        for sub in SECTION_MAP[sk]["subsections"]:
            pool  = [q for q in qs if q["section"] == sub]
            right = sum(q["right"] for q in pool)
            wrong = sum(q["wrong"] for q in pool)
            asked = right + wrong
            sections.append({
                "sub":   sub,
                "name":  SECTION_MAP[sk]["sub_names"].get(sub, ""),
                "total": len(pool),
                "right": right,
                "wrong": wrong,
                "pct":   round(right / asked * 100) if asked > 0 else None,
            })
    total_right = sum(q["right"] for q in qs)
    total_wrong = sum(q["wrong"] for q in qs)
    return jsonify({
        "sections":       sections,
        "total":          len(qs),
        "right":          total_right,
        "wrong":          total_wrong,
        "bookmark_count": len(bookmarks),
    })


@app.route("/api/questions")
def get_questions():
    mode     = request.args.get("mode", "S")
    subs     = request.args.getlist("subs")
    count    = request.args.get("count", type=int)
    adaptive = request.args.get("adaptive", "1") == "1"

    all_qs    = load_all_questions()
    bookmarks = load_bookmarks()

    if mode == "F":
        result = []
        for sk, n in EXAM_DISTRIBUTION.items():
            pool = [q for q in all_qs if q["section"] in SECTION_MAP[sk]["subsections"]]
            if not pool:
                continue
            sampled = weighted_sample(pool, n) if adaptive else random.sample(pool, min(n, len(pool)))
            result.extend(sampled)
        random.shuffle(result)

    elif mode == "W":
        result = [q for q in all_qs if q["wrong"] > 0]
        random.shuffle(result)
        if count:
            result = result[:count]

    elif mode == "B":
        result = [q for q in all_qs if q["row"] in bookmarks]
        random.shuffle(result)
        if count:
            result = result[:count]

    else:  # Study mode
        if not subs:
            return jsonify({"error": "No subsections selected"}), 400
        pool = [q for q in all_qs if q["section"] in subs]
        n    = min(count, len(pool)) if count else len(pool)
        result = weighted_sample(pool, n) if adaptive else random.sample(pool, n)

    return jsonify([format_question(q, bookmarks) for q in result])


@app.route("/api/questions_by_rows", methods=["POST"])
def questions_by_rows():
    data = request.json
    if not data or "rows" not in data:
        return jsonify({"error": "Missing rows"}), 400
    wanted    = set(data["rows"])
    all_qs    = load_all_questions()
    bookmarks = load_bookmarks()
    result    = [q for q in all_qs if q["row"] in wanted]
    random.shuffle(result)
    return jsonify([format_question(q, bookmarks) for q in result])


@app.route("/api/answer", methods=["POST"])
def submit_answer():
    data = request.json
    if not data:
        return jsonify({"error": "No data"}), 400
    row     = data.get("row")
    correct = data.get("correct")
    if not isinstance(row, int) or not isinstance(correct, bool):
        return jsonify({"error": "Invalid payload"}), 400
    save_result(row, correct)
    return jsonify({"ok": True})


@app.route("/api/bookmark", methods=["POST"])
def toggle_bookmark():
    data = request.json
    if not data or "row" not in data:
        return jsonify({"error": "Missing row"}), 400
    row       = data["row"]
    bookmarks = load_bookmarks()
    if row in bookmarks:
        bookmarks.discard(row)
        state = False
    else:
        bookmarks.add(row)
        state = True
    save_bookmarks(bookmarks)
    return jsonify({"bookmarked": state})


@app.route("/api/log", methods=["POST"])
def log_session():
    data = request.json or {}
    append_session_log(
        data.get("mode", "Web"),
        data.get("sections", ""),
        data.get("target", 0),
        data.get("correct", 0),
        data.get("total", 0),
    )
    return jsonify({"ok": True})


@app.route("/api/history")
def history():
    if not os.path.exists(LOG_PATH):
        return jsonify([])
    try:
        with open(LOG_PATH, "r", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        return jsonify(rows[-30:])
    except Exception:
        return jsonify([])


if __name__ == "__main__":
    print("\n  SIE Quizzer Web App")
    print("  Open http://localhost:5000 in your browser\n")
    app.run(debug=False, port=5000)
